"""
scripts/cargar_catalogos.py
---------------------------
Carga inicial de los catálogos a Firestore. Se ejecuta UNA SOLA VEZ, desde
tu computadora (no en Render), antes de empezar a usar el sistema.

Qué hace:
1. Lee el Excel original (catalogos/Datos_para_doc_de_incidencias.xlsx).
2. Sube los 81 clientes a la colección 'clientes', normalizando las
   frecuencias y marcando los clientes con cobro por recolección
   (Walmart, Bodega Aurrera y Sam's).
3. Sube los 8 tipos de incidencia definidos por las PM, con su gravedad
   sugerida del semáforo.

Uso (desde la carpeta backend/):

    # 1. Apunta a tus credenciales de Firebase
    export GOOGLE_APPLICATION_CREDENTIALS=/ruta/a/serviceAccount.json
    #    (en Windows PowerShell:  $env:GOOGLE_APPLICATION_CREDENTIALS="C:\ruta\serviceAccount.json")

    # 2. Ejecuta la carga
    python scripts/cargar_catalogos.py

El script es seguro de re-ejecutar: no duplica registros, porque verifica
si el nombre ya existe antes de crear cada documento.
"""

import os
import sys

from openpyxl import load_workbook

# Permite importar los módulos del backend al ejecutar el script directamente
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modulos import firebase_db as fdb          # noqa: E402
from modulos import utilidades as ut            # noqa: E402

RUTA_EXCEL = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "catalogos", "Datos_para_doc_de_incidencias.xlsx",
)

# Clientes a los que se cobra por recolección: una segunda vuelta por
# material olvidado NO se les puede cobrar (regla de negocio de las PM).
PREFIJOS_COBRO = ("walmart", "bodega", "sam")

TIPOS_INCIDENCIA = [
    ("Modificación de instrucciones",
     "Cambios de indicaciones sin notificación a PM.", "Naranja"),
    ("Problemas de conducta / Servicio",
     "Mala actitud o faltas al código de conducta.", "Naranja"),
    ("Error en información operativa",
     "Los recolectores dan indicaciones o información errónea al cliente.", "Amarillo"),
    ("Material olvidado",
     "No se llevó material solicitado (contenedores, bitácoras, costales).", "Naranja"),
    ("Desvío de ruta / Vueltas excedentes",
     "Vueltas extra por material dejado en sitio u olvidos.", "Naranja"),
    ("Incumplimiento de horario",
     "Llegadas tarde o anticipadas sin programación.", "Amarillo"),
    ("Robo de material",
     "Faltante de material por robo.", "Rojo"),
    ("Olvido de bitácoras o báscula",
     "No se lleva báscula o bitácoras; el registro se hace en bodega.", "Verde"),
]


def leer_clientes_del_excel():
    """Extrae los clientes de la hoja 'Datos generales' del Excel."""
    wb = load_workbook(RUTA_EXCEL, data_only=True)

    hoja = None
    for ws in wb.worksheets:
        if "generales" in ut.sin_acentos_minusculas(ws.title):
            hoja = ws
            break
    if hoja is None:
        raise RuntimeError("No se encontró la hoja 'Datos generales' en el Excel.")

    # Localizar la fila de encabezados
    fila_encabezado = None
    for fila in hoja.iter_rows(min_col=1, max_col=1):
        valor = ut.sin_acentos_minusculas(fila[0].value or "")
        if valor.startswith("nombre del cliente"):
            fila_encabezado = fila[0].row
            break
    if fila_encabezado is None:
        raise RuntimeError("No se encontró el encabezado 'Nombre del Cliente'.")

    clientes = []
    for fila in hoja.iter_rows(min_row=fila_encabezado + 1, max_col=4, values_only=True):
        nombre = ut.limpiar_texto(fila[0])
        if not nombre:
            continue

        estado = ut.limpiar_texto(fila[2]) or "Ciudad de México"
        if ut.sin_acentos_minusculas(estado) == "estado de mexico":
            estado = "Estado de México"

        frecuencia, activo = ut.normalizar_frecuencia(fila[3])
        clientes.append({
            "nombre": nombre,
            "categoria": ut.limpiar_texto(fila[1]) or "Empresa",
            "estado": estado,
            "frecuencia": frecuencia,
            "frecuencia_original": ut.limpiar_texto(fila[3]),
            "activo": activo,
            "cobro_por_recoleccion":
                ut.sin_acentos_minusculas(nombre).startswith(PREFIJOS_COBRO),
            "estatus": "Activo" if activo else "Sin recolecciones",
        })
    return clientes


def main():
    print("Conectando con Firebase...")
    fdb.inicializar()
    base = fdb.db()

    # --- Tipos de incidencia -----------------------------------------------
    print("\nCargando tipos de incidencia...")
    existentes = {t["nombre"] for t in fdb.listar_tipos()}
    nuevos = 0
    for nombre, descripcion, gravedad in TIPOS_INCIDENCIA:
        if nombre in existentes:
            print(f"  · Ya existe: {nombre}")
            continue
        fdb.crear_tipo({"nombre": nombre, "descripcion": descripcion,
                        "gravedad_default": gravedad})
        nuevos += 1
        print(f"  ✓ {nombre} ({gravedad})")
    print(f"Tipos nuevos: {nuevos}")

    # --- Clientes -----------------------------------------------------------
    print("\nCargando clientes desde el Excel...")
    clientes = leer_clientes_del_excel()
    existentes = {c["nombre"] for c in fdb.listar_clientes()}

    # Se usa un lote (batch) para subir muchos documentos de golpe:
    # es mucho más rápido y consume menos cuota de escritura.
    lote = base.batch()
    nuevos, en_lote = 0, 0
    for cliente in clientes:
        if cliente["nombre"] in existentes:
            continue
        ref = base.collection(fdb.COL_CLIENTES).document()
        lote.set(ref, cliente)
        nuevos += 1
        en_lote += 1
        if en_lote >= 400:          # Firestore permite hasta 500 por lote
            lote.commit()
            lote = base.batch()
            en_lote = 0
    if en_lote:
        lote.commit()

    print(f"Clientes en el Excel: {len(clientes)}")
    print(f"Clientes nuevos subidos: {nuevos}")

    # --- Resumen ------------------------------------------------------------
    todos = fdb.listar_clientes()
    print("\n--- Resumen en Firestore ---")
    print(f"Total de clientes: {len(todos)}")
    print(f"Total de tipos de incidencia: {len(fdb.listar_tipos())}")
    print(f"Con cobro por recolección: "
          f"{sum(1 for c in todos if c.get('cobro_por_recoleccion'))}")
    print(f"Sin recolecciones activas: "
          f"{sum(1 for c in todos if not c.get('activo'))}")
    print("\n¡Listo! Los catálogos ya están en Firestore.")


if __name__ == "__main__":
    main()
