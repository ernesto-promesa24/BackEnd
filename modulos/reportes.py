"""
reportes.py
-----------
Generación de reportes en Excel (OpenPyXL) a partir de los documentos de
Firestore. El archivo se construye EN MEMORIA y se envía al navegador: no
se escribe nada en el disco de Render (que es efímero).

Hojas del reporte:
1. Resumen      - indicadores obligatorios de las PM.
2. Servicios    - detalle de servicios del periodo.
3. Incidencias  - detalle, con la gravedad coloreada (semáforo).
4. Evidencias   - imágenes incrustadas, descargadas desde Firebase Storage.
"""

import io
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modulos import utilidades as ut

RELLENO_GRAVEDAD = {
    "Rojo": "D32F2F", "Naranja": "F57C00",
    "Amarillo": "FBC02D", "Verde": "388E3C",
}
FUENTE = "Arial"
AZUL = "1F4E79"

TIPOS_REPORTE = ["Semanal", "Mensual", "Trimestral", "Anual",
                 "Por cliente", "Incidencias pendientes"]


# ---------------------------------------------------------------------------
# Selección del periodo
# ---------------------------------------------------------------------------
def filtrar_por_reporte(servicios, incidencias, tipo, valor):
    """Devuelve (servicios, incidencias, sufijo) según el tipo de reporte."""
    campo = {"Semanal": "semana", "Mensual": "mes",
             "Trimestral": "trimestre", "Anual": "anio"}.get(tipo)

    if tipo == "Incidencias pendientes":
        incidencias_f = [i for i in incidencias if not i.get("resuelta")]
        ids = {i["servicio_id"] for i in incidencias_f}
        return [s for s in servicios if s["id"] in ids], incidencias_f, "pendientes"

    if tipo == "Por cliente":
        servicios_f = [s for s in servicios if s.get("cliente") == valor]
        sufijo = ut.sin_acentos_minusculas(valor).replace(" ", "_")[:30]
    elif campo:
        servicios_f = [s for s in servicios if s.get(campo) == valor]
        sufijo = str(valor)
    else:
        servicios_f, sufijo = servicios, ""

    ids = {s["id"] for s in servicios_f}
    incidencias_f = [i for i in incidencias if i.get("servicio_id") in ids]
    return servicios_f, incidencias_f, sufijo


# ---------------------------------------------------------------------------
# Construcción del Excel
# ---------------------------------------------------------------------------
def generar_excel(titulo, servicios, incidencias, incluir_evidencias=True):
    """Devuelve un BytesIO con el archivo .xlsx listo para descargar."""
    wb = Workbook()
    _hoja_resumen(wb.active, titulo, servicios, incidencias)
    _hoja_servicios(wb.create_sheet("Servicios"), servicios)
    _hoja_incidencias(wb.create_sheet("Incidencias"), incidencias)
    if incluir_evidencias:
        _hoja_evidencias(wb.create_sheet("Evidencias"), incidencias)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ----------------------------- Estilos -------------------------------------
def _encabezado(celda):
    celda.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)
    celda.fill = PatternFill("solid", fgColor=AZUL)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = Border(*(Side(style="thin"),) * 4)


def _tabla(ws, columnas, filas):
    """columnas: [(clave, 'Título')...]  filas: lista de dicts."""
    for col, (_, titulo) in enumerate(columnas, start=1):
        _encabezado(ws.cell(row=1, column=col, value=titulo))

    for fila_num, registro in enumerate(filas, start=2):
        for col, (clave, _) in enumerate(columnas, start=1):
            valor = registro.get(clave)
            if isinstance(valor, bool):
                valor = "Sí" if valor else "No"
            elif valor is None:
                valor = ""
            celda = ws.cell(row=fila_num, column=col, value=valor)
            celda.font = Font(name=FUENTE, size=10)
            celda.border = Border(*(Side(style="thin"),) * 4)

    for col, (clave, titulo) in enumerate(columnas, start=1):
        largos = [len(str(titulo))] + [len(str(r.get(clave) or "")) for r in filas]
        ws.column_dimensions[get_column_letter(col)].width = min(45, max(12, max(largos) + 2))


def _desglose_por_periodo(servicios, incidencias, campo):
    """Retrasos y vueltas por periodo ('mes' o 'trimestre'), con su %."""
    periodos = sorted({s.get(campo) for s in servicios if s.get(campo)})
    filas = []
    for p in periodos:
        serv_p = [s for s in servicios if s.get(campo) == p]
        inci_p = [i for i in incidencias if i.get(campo) == p]
        total_p = len(serv_p)

        r_horario = sum(1 for s in serv_p
                        if s.get("hora_programada") and s.get("hora_real")
                        and s["hora_real"] > s["hora_programada"])
        r_tipo = sum(1 for i in inci_p if i.get("tipo_incidencia") == ut.TIPO_RETRASO)
        retrasos = max(r_horario, r_tipo)

        v_cap = sum(i.get("vueltas_adicionales") or 0 for i in inci_p)
        v_tipo = sum(1 for i in inci_p if i.get("tipo_incidencia") == ut.TIPO_VUELTAS)
        vueltas = v_cap or v_tipo

        filas.append((
            p, total_p,
            retrasos, f"{(100*retrasos/total_p if total_p else 0):.1f}%",
            vueltas, f"{(100*vueltas/total_p if total_p else 0):.1f}%",
        ))
    return filas


# ----------------------------- Resumen -------------------------------------
def _hoja_resumen(ws, titulo, servicios, incidencias):
    ws.title = "Resumen"
    ws["A1"] = titulo
    ws["A1"].font = Font(name=FUENTE, bold=True, size=16, color=AZUL)
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name=FUENTE, italic=True, size=10)

    total_serv, total_inci = len(servicios), len(incidencias)
    serv_con = len({i["servicio_id"] for i in incidencias})
    pct_con = 100 * serv_con / total_serv if total_serv else 0
    # Retrasos: se toma el mayor entre los detectados por horario (hora real
    # posterior a la programada) y los capturados como incidencia de tipo
    # "Incumplimiento de horario". Mismo criterio que el dashboard.
    retrasos_horario = sum(
        1 for s in servicios
        if s.get("hora_programada") and s.get("hora_real")
        and s["hora_real"] > s["hora_programada"])
    retrasos_tipo = sum(1 for i in incidencias
                        if i.get("tipo_incidencia") == ut.TIPO_RETRASO)
    retrasos = max(retrasos_horario, retrasos_tipo)

    vueltas_capturadas = sum(i.get("vueltas_adicionales") or 0 for i in incidencias)
    vueltas_por_tipo = sum(1 for i in incidencias if i.get("tipo_incidencia") == ut.TIPO_VUELTAS)
    vueltas = vueltas_capturadas or vueltas_por_tipo
    quejas = sum(1 for i in incidencias if i.get("reporto_cliente"))
    abiertas = sum(1 for i in incidencias if not i.get("resuelta"))

    resueltas = [i for i in incidencias if i.get("resuelta")]
    a_tiempo = sum(1 for i in resueltas if i.get("resuelta_a_tiempo"))
    pct_a_tiempo = 100 * a_tiempo / len(resueltas) if resueltas else 0

    inci_prioritarios = sum(1 for i in incidencias if i.get("es_prioritario"))
    segundas_vueltas = sum(1 for i in incidencias if i.get("segunda_vuelta_no_cobrable"))

    indicadores = [
        ("Total de servicios", total_serv),
        ("Total de incidencias", total_inci),
        ("% de servicios con incidencia", f"{pct_con:.1f}%"),
        ("% de recolecciones sin incidencias", f"{(100-pct_con if total_serv else 0):.1f}%"),
        ("Retrasos (incumplimiento de horario)", retrasos),
        ("% de retrasos sobre servicios", f"{(100*retrasos/total_serv if total_serv else 0):.1f}%"),
        ("Vueltas extra", vueltas),
        ("% de vueltas extra sobre servicios", f"{(100*vueltas/total_serv if total_serv else 0):.1f}%"),
        ("Reportes por parte del cliente (quejas)", quejas),
        ("% de quejas sobre incidencias", f"{(100*quejas/total_inci if total_inci else 0):.1f}%"),
        ("% de incidencias resueltas a tiempo", f"{pct_a_tiempo:.1f}%"),
        ("Incidencias abiertas", abiertas),
        ("Incidencias cerradas", total_inci - abiertas),
        ("Incidencias de clientes prioritarios", inci_prioritarios),
        ("2ª vueltas no cobrables (Walmart/Bodega/Sam's)", segundas_vueltas),
    ]

    fila = 4
    _encabezado(ws.cell(row=fila, column=1, value="Indicador"))
    _encabezado(ws.cell(row=fila, column=2, value="Valor"))
    for nombre, valor in indicadores:
        fila += 1
        c1 = ws.cell(row=fila, column=1, value=nombre)
        c2 = ws.cell(row=fila, column=2, value=valor)
        c1.font = Font(name=FUENTE, size=10)
        c2.font = Font(name=FUENTE, size=10, bold=True)
        c1.border = c2.border = Border(*(Side(style="thin"),) * 4)

    # % por tipo de incidencia (indicador obligatorio)
    if total_inci:
        from collections import Counter
        fila += 2
        celda = ws.cell(row=fila, column=1, value="Porcentaje por tipo de incidencia")
        celda.font = Font(name=FUENTE, bold=True, size=12, color=AZUL)
        fila += 1
        for col, tit in enumerate(["Tipo de incidencia", "Cantidad", "%"], start=1):
            _encabezado(ws.cell(row=fila, column=col, value=tit))
        for tipo, cantidad in Counter(i["tipo_incidencia"] for i in incidencias).most_common():
            fila += 1
            ws.cell(row=fila, column=1, value=tipo).font = Font(name=FUENTE, size=10)
            ws.cell(row=fila, column=2, value=cantidad).font = Font(name=FUENTE, size=10)
            ws.cell(row=fila, column=3,
                    value=f"{100*cantidad/total_inci:.1f}%").font = Font(name=FUENTE, size=10)
            for col in (1, 2, 3):
                ws.cell(row=fila, column=col).border = Border(*(Side(style="thin"),) * 4)

    # Desglose mensual y trimestral de retrasos y vueltas (indicador obligatorio)
    for campo, titulo_bloque in [("mes", "Retrasos y vueltas por mes"),
                                 ("trimestre", "Retrasos y vueltas por trimestre")]:
        desglose = _desglose_por_periodo(servicios, incidencias, campo)
        if not desglose:
            continue
        fila += 2
        celda = ws.cell(row=fila, column=1, value=titulo_bloque)
        celda.font = Font(name=FUENTE, bold=True, size=12, color=AZUL)
        fila += 1
        encabezados = ["Periodo", "Servicios", "Retrasos", "% retrasos",
                       "Vueltas", "% vueltas"]
        for col, tit in enumerate(encabezados, start=1):
            _encabezado(ws.cell(row=fila, column=col, value=tit))
        for registro in desglose:
            fila += 1
            for col, valor in enumerate(registro, start=1):
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.font = Font(name=FUENTE, size=10)
                celda.border = Border(*(Side(style="thin"),) * 4)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12


# ----------------------------- Servicios -----------------------------------
def _hoja_servicios(ws, servicios):
    if not servicios:
        ws["A1"] = "Sin servicios en el periodo."
        ws["A1"].font = Font(name=FUENTE, italic=True)
        return
    _tabla(ws, [
        ("fecha", "Fecha"), ("cliente", "Cliente"), ("categoria", "Categoría"),
        ("estado", "Estado"), ("tipo_servicio", "Tipo de servicio"),
        ("hora_programada", "Hora programada"), ("hora_real", "Hora real"),
        ("realizado", "Realizado"), ("resultado", "Resultado"),
        ("responsable", "Responsable"), ("observaciones", "Observaciones"),
    ], servicios)


# ----------------------------- Incidencias ---------------------------------
def _hoja_incidencias(ws, incidencias):
    if not incidencias:
        ws["A1"] = "Sin incidencias en el periodo."
        ws["A1"].font = Font(name=FUENTE, italic=True)
        return

    filas = []
    for i in incidencias:
        copia = dict(i)
        copia["num_evidencias"] = len(i.get("evidencias") or [])
        if copia.get("resuelta_a_tiempo") is None:
            copia["resuelta_a_tiempo"] = "—"
        filas.append(copia)

    columnas = [
        ("fecha", "Fecha"), ("cliente", "Cliente"),
        ("categoria", "Categoría"), ("estado", "Estado"),
        ("frecuencia", "Frecuencia"),
        ("tipo_incidencia", "Tipo de incidencia"), ("gravedad", "Gravedad"),
        ("descripcion", "Descripción"),
        ("vueltas_adicionales", "Vueltas adic."),
        ("minutos_retraso", "Retraso (min)"),
        ("reporto_cliente", "Reportó el cliente"), ("resuelta", "Resuelta"),
        ("resuelta_a_tiempo", "Resuelta a tiempo"),
        ("fecha_resolucion", "Fecha de resolución"),
        ("accion_correctiva", "Acción correctiva"),
        ("es_prioritario", "Prioritario"),
        ("segunda_vuelta_no_cobrable", "2ª vuelta no cobrable"),
        ("material_pendiente", "Material que quedó"),
        ("responsable", "Responsable"), ("num_evidencias", "Evidencias"),
        ("comentarios", "Comentarios"),
    ]
    _tabla(ws, columnas, filas)

    # Colorear la columna de gravedad con el semáforo
    col_gravedad = [c for c, _ in columnas].index("gravedad") + 1
    for fila in range(2, len(filas) + 2):
        celda = ws.cell(row=fila, column=col_gravedad)
        color = RELLENO_GRAVEDAD.get(celda.value)
        if color:
            celda.fill = PatternFill("solid", fgColor=color)
            celda.font = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")


# ----------------------------- Evidencias ----------------------------------
def _hoja_evidencias(ws, incidencias):
    """Descarga las imágenes desde Firebase Storage y las incrusta."""
    ws["A1"] = "Evidencias de incidencias"
    ws["A1"].font = Font(name=FUENTE, bold=True, size=14, color=AZUL)
    ws.column_dimensions["A"].width = 60

    fila, hubo = 3, False
    for inc in incidencias:
        evidencias = inc.get("evidencias") or []
        if not evidencias:
            continue
        hubo = True
        ws.cell(row=fila, column=1,
                value=f"{inc.get('fecha')} | {inc.get('cliente')} | "
                      f"{inc.get('tipo_incidencia')} ({inc.get('gravedad')})")
        ws.cell(row=fila, column=1).font = Font(name=FUENTE, bold=True, size=11)
        fila += 1

        for ev in evidencias:
            if ev.get("tipo") == "imagen" and ev.get("url"):
                try:
                    respuesta = requests.get(ev["url"], timeout=15)
                    respuesta.raise_for_status()
                    imagen = XLImage(io.BytesIO(respuesta.content))
                    if imagen.width > 320:      # Escalar manteniendo proporción
                        factor = 320 / imagen.width
                        imagen.width = 320
                        imagen.height = int(imagen.height * factor)
                    ws.add_image(imagen, f"A{fila}")
                    fila += max(1, int(imagen.height / 19)) + 1
                except Exception:
                    ws.cell(row=fila, column=1,
                            value=f"(No se pudo descargar: {ev.get('nombre')})")
                    fila += 1
            else:
                celda = ws.cell(row=fila, column=1,
                                value=f"Archivo adjunto: {ev.get('nombre')}")
                celda.font = Font(name=FUENTE, italic=True, size=10)
                if ev.get("url"):
                    celda.hyperlink = ev["url"]
                fila += 1
        fila += 1

    if not hubo:
        ws["A3"] = "Las incidencias del periodo no tienen evidencias adjuntas."
        ws["A3"].font = Font(name=FUENTE, italic=True)
